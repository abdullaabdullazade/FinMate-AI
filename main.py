from fastapi import FastAPI, Request, Depends, UploadFile, File, Form, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse, Response, RedirectResponse
from fastapi.exceptions import RequestValidationError
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import func, extract
from datetime import datetime, timedelta, date as date_type, timezone
from typing import Optional
import os
import base64
import asyncio
from dotenv import load_dotenv
from starlette.middleware.sessions import SessionMiddleware

# Load environment variables
load_dotenv()

from database import get_db, init_db, seed_demo_data, reset_demo_data
from models import User, Expense, ChatMessage, XPLog, Wish, Dream
from ai_service import ai_service
from gamification import gamification
from forecast_service import forecast_service
from voice_service import voice_service
from pdf_generator import pdf_generator
import hashlib

# Currency conversion rates to AZN (approx, can be updated)
CURRENCY_RATES = {
    "AZN": 1.0,
    "USD": 1.7,
    "EUR": 1.82,
    "TRY": 0.055,
    "RUB": 0.018,
    "GBP": 2.15
}

# Initialize FastAPI app
app = FastAPI(title="FinMate AI", description="Your Personal CFO Assistant")

# Add session middleware
app.add_middleware(SessionMiddleware, secret_key="finmate-secret-key-change-in-production")

# Templates
templates = Jinja2Templates(directory="templates")

# Create static directory if doesn't exist
os.makedirs("static/uploads", exist_ok=True)
app.mount("/static", StaticFiles(directory="static"), name="static")

# Initialize database on startup
@app.on_event("startup")
async def startup_event():
    init_db()
    seed_demo_data()


# Custom 404 handler
@app.exception_handler(404)
async def custom_404_handler(request: Request, exc):
    """Custom 404 page with crying pet"""
    return templates.TemplateResponse("404.html", {
        "request": request
    }, status_code=404)


# Offline page route
@app.get("/offline", response_class=HTMLResponse)
async def offline_page(request: Request):
    """Offline page when internet connection is lost"""
    return templates.TemplateResponse("offline.html", {
        "request": request
    })


# ==================== HELPER FUNCTIONS ====================

def hash_password(password: str) -> str:
    """Hash password using SHA256"""
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(password: str, password_hash: str) -> bool:
    """Verify password against hash"""
    return hash_password(password) == password_hash

def get_current_user(request: Request, db: Session = Depends(get_db)) -> Optional[User]:
    """Get current logged in user from session"""
    user_id = request.session.get("user_id")
    if not user_id:
        return None
    
    db.expire_all()
    user = db.query(User).filter(User.id == user_id).first()
    
    # Clear session if user doesn't exist (prevents redirect loops)
    if not user:
        request.session.clear()
    
    return user

def require_auth(request: Request, db: Session = Depends(get_db)) -> User:
    """Require authentication, redirect to login if not authenticated"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    return user


def build_db_context(db: Session, user_id: int) -> dict:
    """Build financial context from database for AI"""
    
    # Get current month's expenses
    now = datetime.utcnow()
    month_start = datetime(now.year, now.month, 1)
    
    expenses = db.query(Expense).filter(
        Expense.user_id == user_id,
        Expense.date >= month_start
    ).all()
    
    # Calculate totals
    total_spending = sum(exp.amount for exp in expenses)
    
    # Category breakdown
    category_breakdown = {}
    for exp in expenses:
        if exp.category not in category_breakdown:
            category_breakdown[exp.category] = 0
        category_breakdown[exp.category] += exp.amount
    
    # Sort categories by amount
    category_breakdown = dict(sorted(category_breakdown.items(), key=lambda x: x[1], reverse=True))
    
    # Subscription count
    subscription_count = db.query(Expense).filter(
        Expense.user_id == user_id,
        Expense.is_subscription == True
    ).count()
    
    # Recent expenses (last 5)
    recent_expenses = db.query(Expense).filter(
        Expense.user_id == user_id
    ).order_by(Expense.date.desc()).limit(5).all()
    
    recent_exp_list = [
        {"merchant": exp.merchant, "amount": exp.amount, "category": exp.category}
        for exp in recent_expenses
    ]
    
    # Largest expense this month
    largest_expense = db.query(Expense).filter(
        Expense.user_id == user_id,
        Expense.date >= month_start
    ).order_by(Expense.amount.desc()).first()
    
    largest_exp_dict = None
    if largest_expense:
        largest_exp_dict = {
            "merchant": largest_expense.merchant,
            "amount": largest_expense.amount,
            "category": largest_expense.category
        }
    
    # Get user budget
    user = db.query(User).filter(User.id == user_id).first()
    
    return {
        "total_spending": total_spending,
        "budget": user.monthly_budget if user else 2000.0,
        "category_breakdown": category_breakdown,
        "subscription_count": subscription_count,
        "recent_expenses": recent_exp_list,
        "largest_expense": largest_exp_dict
    }


def convert_currency(amount: float, from_currency: str, to_currency: str) -> float:
    """Convert amount between currencies using static rates"""
    from_rate = CURRENCY_RATES.get(from_currency.upper())
    to_rate = CURRENCY_RATES.get(to_currency.upper())
    if not from_rate or not to_rate or to_currency.upper() == from_currency.upper():
        return amount
    # Normalize to AZN then to target
    azn_value = amount * from_rate if from_currency.upper() != "AZN" else amount
    if to_currency.upper() == "AZN":
        return azn_value
    return azn_value / to_rate




def pseudo_coords_for_merchant(merchant: str) -> tuple:
    """Generate stable pseudo-random Baku coordinates for a merchant"""
    seed = int(hashlib.sha256(merchant.encode("utf-8")).hexdigest(), 16)
    # Baku bounding box approx
    lat = 40.35 + (seed % 1000) / 10000.0  # 40.35 - 40.45
    lon = 49.80 + (seed // 1000 % 1500) / 10000.0  # 49.80 - 49.95
    return (round(lat, 5), round(lon, 5))


def pseudo_coords_for_merchant(merchant: str) -> tuple:
    """Generate pseudo-random but stable Baku coords for a merchant"""
    seed = int(hashlib.sha256(merchant.encode("utf-8")).hexdigest(), 16)
    # Baku rough box: lat 40.35 - 40.45, lon 49.80 - 49.95
    lat = 40.35 + (seed % 1000) / 10000.0
    lon = 49.80 + (seed // 1000 % 1500) / 10000.0
    return (round(lat, 5), round(lon, 5))


def calculate_eco_score(expenses):
    """Calculate total CO2 impact from expenses"""
    total_co2 = 0
    for exp in expenses:
        category = exp.category.lower()
        amount = exp.amount
        
        # CO2 estimation (kg per AZN spent)
        if 'nəqliyyat' in category or 'transport' in category:
            total_co2 += amount * 0.5  # High impact
        elif 'market' in category or 'restoran' in category or 'kafe' in category:
            total_co2 += amount * 0.2  # Medium impact
        else:
            total_co2 += amount * 0.05  # Low impact
    
    # Determine icon based on total
    if total_co2 > 100:
        icon = "🍂"  # Red leaf
    elif total_co2 > 50:
        icon = "🌿"  # Yellow/green leaf
    else:
        icon = "🌱"  # Green leaf
    
    return {
        "value": round(total_co2, 1),
        "icon": icon
    }


def calculate_eco_breakdown(category_data):
    """Calculate CO2 breakdown by category"""
    breakdown = {}
    
    for category, amount in category_data.items():
        cat_lower = category.lower()
        
        if 'nəqliyyat' in cat_lower or 'transport' in cat_lower:
            co2 = amount * 0.5
            icon = "🚗"
            level = "Yüksək təsir"
        elif 'market' in cat_lower or 'restoran' in cat_lower or 'kafe' in cat_lower:
            co2 = amount * 0.2
            icon = "🍔"
            level = "Orta təsir"
        else:
            co2 = amount * 0.05
            icon = "💻"
            level = "Aşağı təsir"
        
        breakdown[category] = {
            "co2": round(co2, 1),
            "icon": icon,
            "level": level
        }
    
    # Sort by CO2 impact
    breakdown = dict(sorted(breakdown.items(), key=lambda x: x[1]['co2'], reverse=True)[:5])
    return breakdown


def get_eco_tip(eco_breakdown):
    """Generate AI eco tip based on spending"""
    tips = [
        "Piyada getməklə 10kg CO₂ qənaət etdiniz! 🚶",
        "Ev yeməyi hazırlamaqla planetə kömək edirsiniz! 🌍",
        "Digital xidmətlər daha az karbon izi buraxır! 💚",
        "İctimai nəqliyyatdan istifadə edin - daha ekolojikdir! 🚌",
        "Yerli məhsullar almaqla daşınma emissiyalarını azaldın! 🥬"
    ]
    import random
    return random.choice(tips)


def detect_financial_personality(user_id: int, db: Session) -> dict:
    """Detect user's financial personality based on spending habits"""
    # Get current month's expenses
    now = datetime.utcnow()
    month_start = datetime(now.year, now.month, 1)
    
    expenses = db.query(Expense).filter(
        Expense.user_id == user_id,
        Expense.date >= month_start
    ).all()
    
    if not expenses:
        return {
            "title": "The Beginner",
            "emoji": "🌱",
            "quote": "Hər böyük səyahət bir addımla başlayır",
            "top_category": "Yeni",
            "spending_score": 5
        }
    
    # Calculate category totals
    category_totals = {}
    total_spending = 0
    for exp in expenses:
        total_spending += exp.amount
        if exp.category not in category_totals:
            category_totals[exp.category] = 0
        category_totals[exp.category] += exp.amount
    
    # Find top category
    top_category = max(category_totals, key=category_totals.get) if category_totals else "Digər"
    top_amount = category_totals.get(top_category, 0)
    top_percentage = (top_amount / total_spending * 100) if total_spending > 0 else 0
    
    # Get user budget
    user = db.query(User).filter(User.id == user_id).first()
    savings_percentage = ((user.monthly_budget - total_spending) / user.monthly_budget * 100) if user.monthly_budget > 0 else 0
    
    # Determine personality
    if savings_percentage > 30:
        return {
            "title": "The Monk",
            "emoji": "🧘",
            "quote": "Maliyyə Zen ustadı",
            "top_category": top_category,
            "spending_score": 9
        }
    elif 'Restoran' in top_category or 'Kafe' in top_category or 'Market' in top_category:
        return {
            "title": "The Foodie",
            "emoji": "🍔",
            "quote": "Dad hər şeydir səninçün",
            "top_category": top_category,
            "spending_score": 7
        }
    elif 'Nəqliyyat' in top_category or 'Transport' in top_category:
        return {
            "title": "The Explorer",
            "emoji": "✈️",
            "quote": "Həmişə hərəkətdəsən",
            "top_category": top_category,
            "spending_score": 6
        }
    elif top_percentage > 40:
        return {
            "title": "The Specialist",
            "emoji": "🎯",
            "quote": "Fokuslanmış və məqsədyönlü",
            "top_category": top_category,
            "spending_score": 8
        }
    else:
        return {
            "title": "The Balanced Pro",
            "emoji": "⚖️",
            "quote": "Tarazlıq sənin gücündür",
            "top_category": top_category,
            "spending_score": 8
        }


# ==================== AUTH ROUTES ====================

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request, db: Session = Depends(get_db)):
    """Login page"""
    # If already logged in and user exists, redirect to dashboard
    user = get_current_user(request, db)
    if user:
        return RedirectResponse(url="/", status_code=303)
    
    # Check if user just registered
    registered = request.query_params.get("registered") == "1"
    
    return templates.TemplateResponse("login.html", {
        "request": request,
        "registered": registered
    })

@app.post("/api/login")
async def login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    """Handle login"""
    user = db.query(User).filter(User.username == username).first()
    
    # Check for demo user (password: "demo")
    if username == "demo" and password == "demo":
        # Find or create demo user
        user = db.query(User).filter(User.username == "demo").first()
        if not user:
            user = User(username="demo", monthly_budget=3000.0, password_hash=None)
            db.add(user)
            db.commit()
            db.refresh(user)
        request.session["user_id"] = user.id
        return RedirectResponse(url="/", status_code=303)
    
    # Check for regular users
    if not user or not user.password_hash:
        return templates.TemplateResponse("login.html", {
            "request": request,
            "error": "İstifadəçi adı və ya şifrə yanlışdır"
        })
    
    if not verify_password(password, user.password_hash):
        return templates.TemplateResponse("login.html", {
            "request": request,
            "error": "İstifadəçi adı və ya şifrə yanlışdır"
        })
    
    request.session["user_id"] = user.id
    return RedirectResponse(url="/", status_code=303)

@app.get("/signup", response_class=HTMLResponse)
async def signup_page(request: Request, db: Session = Depends(get_db)):
    """Signup page"""
    # If already logged in and user exists, redirect to dashboard
    user = get_current_user(request, db)
    if user:
        return RedirectResponse(url="/", status_code=303)
    return templates.TemplateResponse("signup.html", {"request": request})

@app.post("/api/signup")
async def signup(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    confirm_password: str = Form(...),
    db: Session = Depends(get_db)
):
    """Handle signup"""
    # Validation
    if password != confirm_password:
        return templates.TemplateResponse("signup.html", {
            "request": request,
            "error": "Şifrələr uyğun gəlmir"
        })
    
    if len(password) < 4:
        return templates.TemplateResponse("signup.html", {
            "request": request,
            "error": "Şifrə ən azı 4 simvol olmalıdır"
        })
    
    if len(username) < 3:
        return templates.TemplateResponse("signup.html", {
            "request": request,
            "error": "İstifadəçi adı ən azı 3 simvol olmalıdır"
        })
    
    # Check if user exists
    existing_user = db.query(User).filter(User.username == username).first()
    if existing_user:
        return templates.TemplateResponse("signup.html", {
            "request": request,
            "error": "Bu istifadəçi adı artıq mövcuddur"
        })
    
    # Create new user (monthly_income will be set during onboarding after login)
    password_hash = hash_password(password)
    new_user = User(
        username=username,
        password_hash=password_hash,
        monthly_budget=1000.0
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    
    # Redirect to login page with success message
    return RedirectResponse(url="/login?registered=1", status_code=303)

@app.get("/logout")
async def logout(request: Request):
    """Logout user"""
    request.session.clear()
    return RedirectResponse(url="/login", status_code=303)

# ==================== ROUTES ====================

@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request, db: Session = Depends(get_db)):
    """Main dashboard page"""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    # Greeting
    hour = datetime.now().hour
    if hour < 12:
        greeting = "Sabahın xeyir"
    elif hour < 18:
        greeting = "Günün aydın"
    else:
        greeting = "Axşamın xeyir"
    
    # Track login streak
    today = date_type.today()
    if user.last_login_date != today:
        if user.last_login_date == today - timedelta(days=1):
            # Consecutive day login
            user.login_streak += 1
            gamification.award_xp(user, "daily_login", db)
        else:
            # Streak broken
            user.login_streak = 1
        user.last_login_date = today
        db.commit()
    
    # Get current month's expenses
    now = datetime.utcnow()
    month_start = datetime(now.year, now.month, 1)
    
    expenses = db.query(Expense).filter(
        Expense.user_id == user.id,
        Expense.date >= month_start
    ).all()
    
    # Calculate stats
    total_spending = sum(exp.amount for exp in expenses)
    
    # Category breakdown
    category_data = {}
    for exp in expenses:
        if exp.category not in category_data:
            category_data[exp.category] = 0
        category_data[exp.category] += exp.amount
    
    # Recent expenses (last 10 by creation time so freshly scanned items appear)
    recent_expenses = db.query(Expense).filter(
        Expense.user_id == user.id
    ).order_by(Expense.created_at.desc()).limit(10).all()
    
    # Budget percentage
    budget_percentage = (total_spending / user.monthly_budget * 100) if user.monthly_budget > 0 else 0
    
    # Check daily budget limit
    daily_limit_alert = None
    if user.daily_budget_limit:
        today = date_type.today()
        today_expenses = db.query(Expense).filter(
            Expense.user_id == user.id,
            Expense.date >= today,
            Expense.date < today + timedelta(days=1)
        ).all()
        today_total = sum(exp.amount for exp in today_expenses)
        
        if today_total > user.daily_budget_limit:
            daily_limit_alert = {
                "exceeded": True,
                "today_spending": today_total,
                "limit": user.daily_budget_limit,
                "over_by": today_total - user.daily_budget_limit
            }
        elif today_total >= user.daily_budget_limit * 0.9:
            daily_limit_alert = {
                "exceeded": False,
                "warning": True,
                "today_spending": today_total,
                "limit": user.daily_budget_limit,
                "remaining": user.daily_budget_limit - today_total
            }
    
    # Get level info for gamification
    level_info = gamification.get_level_info(user.xp_points)
    
    # Get forecast data
    forecast = forecast_service.get_forecast(user.id, db)
    # Wishlist items
    wishes = db.query(Wish).filter(Wish.user_id == user.id).order_by(Wish.created_at.desc()).limit(10).all()

    # ===== NEW: ECO-SCORE CALCULATION =====
    eco_score = calculate_eco_score(expenses)
    eco_breakdown = calculate_eco_breakdown(category_data)
    eco_tip = get_eco_tip(eco_breakdown)

    # ===== NEW: DREAM VAULT DATA =====
    dream_goal = 2999.0  # PS5 price
    dream_saved = 450.0  # Demo data - in production, store in DB
    dream_progress = (dream_saved / dream_goal) * 100
    dream_blur = max(0, 20 - (dream_progress / 100 * 20))  # 20px at 0%, 0px at 100%

    # ===== NEW: LOCAL GEMS - Find expensive expenses and suggest alternatives =====
    from local_gems import find_local_gems, format_gem_suggestion
    local_gems_suggestions = []
    
    # Find expensive expenses (above average or specific expensive merchants)
    expensive_merchants = ["Starbucks", "McDonald's", "KFC", "Papa John's", "Kino", "Cinema"]
    for expense in recent_expenses[:5]:  # Check last 5 expenses
        merchant = expense.merchant
        amount = expense.amount
        
        # Check if it's an expensive merchant or above average amount
        is_expensive = any(exp_merchant.lower() in merchant.lower() for exp_merchant in expensive_merchants)
        avg_amount = total_spending / len(expenses) if expenses else 0
        is_above_avg = amount > avg_amount * 1.5 if avg_amount > 0 else False
        
        if is_expensive or is_above_avg:
            alternatives = find_local_gems(merchant, amount, expense.category)
            if alternatives:
                gem_text = format_gem_suggestion(merchant, amount, alternatives)
                if gem_text:
                    local_gems_suggestions.append({
                        "merchant": merchant,
                        "amount": amount,
                        "category": expense.category,
                        "suggestions": gem_text,
                        "alternatives": alternatives[:2]  # Top 2 alternatives
                    })
    
    # Limit to 3 suggestions max
    local_gems_suggestions = local_gems_suggestions[:3]

    # Check if user needs budget onboarding (budget is default 1000)
    show_income_onboarding = user.monthly_budget == 1000.0 and (not hasattr(user, 'monthly_income') or getattr(user, 'monthly_income', None) is None)
    
    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "user": user,
        "total_spending": total_spending,
        "budget_percentage": budget_percentage,
        "category_data": category_data,
        "recent_expenses": recent_expenses,
        "daily_limit_alert": daily_limit_alert,
        "level_info": level_info,
        "forecast": forecast,
        "wishes": wishes,
        "min": min,
        "float": float,
        "now": now,
        "greeting": greeting,
        # New data
        "eco_score": eco_score,
        "eco_breakdown": eco_breakdown,
        "eco_tip": eco_tip,
        "dream_goal": dream_goal,
        "dream_saved": dream_saved,
        "dream_progress": dream_progress,
        "dream_blur": dream_blur,
        "local_gems": local_gems_suggestions,
        "daily_limit_alert": daily_limit_alert,
        "show_income_onboarding": show_income_onboarding
    })


@app.get("/chat", response_class=HTMLResponse)
async def chat_page(request: Request, db: Session = Depends(get_db)):
    """AI Chat page"""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    # Get chat history
    messages = db.query(ChatMessage).filter(
        ChatMessage.user_id == user.id
    ).order_by(ChatMessage.timestamp.asc()).all()
    
    return templates.TemplateResponse("chat.html", {
        "request": request,
        "user": user,
        "messages": messages
    })


@app.post("/api/chat")
async def send_chat_message(
    request: Request,
    message: str = Form(...),
    db: Session = Depends(get_db)
):
    """Handle chat message from user"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    # Save user message
    user_msg = ChatMessage(
        user_id=user.id,
        role="user",
        content=message,
        timestamp=datetime.utcnow()
    )
    db.add(user_msg)
    db.commit()
    
    # Build database context
    db_context = build_db_context(db, user.id)
    
    # Get recent chat history
    recent_messages = db.query(ChatMessage).filter(
        ChatMessage.user_id == user.id
    ).order_by(ChatMessage.timestamp.desc()).limit(10).all()
    
    chat_history = [
        {"role": msg.role, "content": msg.content}
        for msg in reversed(recent_messages)
    ]
    
    # Get AI response with dynamic persona (force Azerbaijani replies)
    ai_response = ai_service.chat_with_cfo(
        message,
        db_context,
        chat_history,
        "az",
        user  # Pass user for behavioral profiling
    )
    
    # Convert markdown to HTML in AI response for better formatting
    import re
    
    # Bold text: **text** -> <strong>text</strong>
    ai_response_formatted = re.sub(r'\*\*(.*?)\*\*', r'<strong>\1</strong>', ai_response)
    
    # Italic text: *text* (but not ** which is bold)
    ai_response_formatted = re.sub(r'(?<!\*)\*(?!\*)([^\*]+?)\*(?!\*)', r'<em>\1</em>', ai_response_formatted)
    
    # Bullet points: ↑ or • at start of line -> styled bullet
    ai_response_formatted = re.sub(r'(^|\n)([↑•])\s*(.+)', r'\1<span class="chat-bullet">\2</span> \3', ai_response_formatted)
    
    # Line breaks: convert \n to <br> for proper display
    ai_response_formatted = ai_response_formatted.replace('\n', '<br>')
    
    # Save AI response
    ai_msg = ChatMessage(
        user_id=user.id,
        role="ai",
        content=ai_response_formatted,
        timestamp=datetime.utcnow()
    )
    db.add(ai_msg)
    db.commit()
    
    # Award XP for chat interaction
    xp_result = gamification.award_xp(user, "chat_message", db)
    db.refresh(user)  # Refresh to get updated XP
    
    # If client already appended user message, only return AI bubble
    client_appended = request.headers.get("X-Client-Appended") == "true"
    outgoing_messages = [ai_msg] if client_appended else [user_msg, ai_msg]
    
    # Return HTML for HTMX
    return templates.TemplateResponse("partials/chat_messages.html", {
        "request": request,
        "messages": outgoing_messages,
        "xp_result": xp_result
    })


@app.get("/scan", response_class=HTMLResponse)
async def scan_page(request: Request, db: Session = Depends(get_db)):
    """Receipt scanner page"""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("scan.html", {
        "request": request,
        "user": user
    })


@app.post("/api/scan-receipt")
async def scan_receipt(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Process uploaded receipt image"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    try:
        # Save uploaded file
        file_path = f"static/uploads/{file.filename}"
        with open(file_path, "wb") as f:
            content = await file.read()
            f.write(content)
        
        # Analyze receipt with AI
        receipt_data = ai_service.analyze_receipt(file_path)

        # Normalize payload to avoid JSON serialization issues
        items = receipt_data.get("items", [])
        if not isinstance(items, list):
            try:
                items = list(items)
            except Exception:
                items = []
        receipt_data["items"] = items
        receipt_data["merchant"] = receipt_data.get("merchant") or "Unknown Merchant"
        receipt_data["suggested_category"] = receipt_data.get("suggested_category") or "Other"
        # Don't override date if AI extracted it, only set default if missing
        if not receipt_data.get("date") or receipt_data.get("date") == "null":
            # Will use current date/time from browser or UTC+4
            receipt_data["date"] = None
    
        # Check if this is actually a receipt
        if not receipt_data.get("is_receipt", True):
            return templates.TemplateResponse("partials/receipt_result.html", {
                "request": request,
                "receipt_data": {
                    "error": "Bu qəbz deyil. Xahiş edirik qəbz şəkli yükləyin.",
                    "is_not_receipt": True
                },
                "success": False
            })
        
        original_currency = receipt_data.get("currency", "AZN").upper() if isinstance(receipt_data, dict) else "AZN"
        conversion_note = None
        if original_currency != "AZN" and original_currency in CURRENCY_RATES:
            rate = CURRENCY_RATES[original_currency]
            try:
                original_total = float(receipt_data.get("total", 0.0))
                converted_total = convert_currency(original_total, original_currency, "AZN")
                receipt_data["total_original"] = original_total
                receipt_data["total"] = round(converted_total, 2)
                receipt_data["original_currency"] = original_currency
                receipt_data["currency"] = "AZN"
                conversion_note = f"{original_currency} → AZN (məzənnə {rate})"

                # Convert items too
                items = receipt_data.get("items", [])
                for item in items:
                    price = float(item.get("price", 0.0))
                    item["price_original"] = price
                    item["price"] = round(convert_currency(price, original_currency, "AZN"), 2)
                receipt_data["items"] = items
            except Exception as conv_err:
                print(f"Currency conversion failed: {conv_err}")

        # Surface non-blocking AI warnings (offline/fallback mode)
        if not conversion_note and isinstance(receipt_data, dict):
            conversion_note = receipt_data.get("note")
    
        # Check if error occurred
        if "error" in receipt_data:
            return templates.TemplateResponse("partials/receipt_result.html", {
                "request": request,
                "receipt_data": receipt_data,
                "success": False
            })

        # Auto-save expense (skip manual confirmation)
        try:
            raw_date = receipt_data.get("date")
            # Get client date/time from request headers (browser's local time)
            client_date_str = request.headers.get("X-Client-Date", None)
            
            try:
                raw_time = receipt_data.get("time")
                if raw_date and raw_date != "null" and raw_date.lower() != "none":
                    # Use date from receipt if available (assume it's in Azerbaijan time UTC+4)
                    if raw_time and raw_time != "null" and raw_time.lower() != "none":
                        # Both date and time from receipt - treat as UTC+4
                        try:
                            time_parts = raw_time.split(":")
                            hour = int(time_parts[0]) if len(time_parts) > 0 else 0
                            minute = int(time_parts[1]) if len(time_parts) > 1 else 0
                            # Receipt time is already in Azerbaijan time (UTC+4), so use it directly
                            expense_date = datetime.strptime(raw_date, "%Y-%m-%d").replace(hour=hour, minute=minute)
                        except:
                            expense_date = datetime.strptime(raw_date, "%Y-%m-%d")
                    else:
                        # Only date from receipt, use current time in UTC+4
                        az_timezone = timezone(timedelta(hours=4))
                        now_az = datetime.now(az_timezone)
                        expense_date = datetime.strptime(raw_date, "%Y-%m-%d").replace(hour=now_az.hour, minute=now_az.minute)
                elif client_date_str:
                    # Use browser's local time, but convert to UTC+4 (Azerbaijan time)
                    from dateutil import parser
                    client_datetime = parser.isoparse(client_date_str)
                    # If browser sends UTC time, convert to UTC+4
                    # If browser sends local time, assume it's already in user's timezone
                    # For Azerbaijan, we want UTC+4
                    if client_datetime.tzinfo is None:
                        # No timezone info, assume it's UTC and convert to UTC+4
                        az_timezone = timezone(timedelta(hours=4))
                        expense_date = client_datetime.replace(tzinfo=timezone.utc).astimezone(az_timezone).replace(tzinfo=None)
                    else:
                        # Has timezone, convert to UTC+4
                        az_timezone = timezone(timedelta(hours=4))
                        expense_date = client_datetime.astimezone(az_timezone).replace(tzinfo=None)
                else:
                    # Fallback to UTC+4 (Azerbaijan time) - current date and time
                    az_timezone = timezone(timedelta(hours=4))
                    expense_date = datetime.now(az_timezone).replace(tzinfo=None)
            except Exception as date_err:
                # Fallback to UTC+4 - current date and time
                print(f"Date parsing error: {date_err}, using current time")
                az_timezone = timezone(timedelta(hours=4))
                expense_date = datetime.now(az_timezone).replace(tzinfo=None)

            items_list = receipt_data.get("items", [])
            if not isinstance(items_list, list):
                items_list = []

            expense = Expense(
                user_id=user.id,
                amount=float(receipt_data.get("total", 0.0) or 0.0),
                merchant=receipt_data.get("merchant") or "Unknown Merchant",
                category=receipt_data.get("suggested_category") or "Other",
                date=expense_date,
                items=items_list
            )
            db.add(expense)
            db.commit()
            db.refresh(expense)
            
            # Update receipt_data with formatted date/time for display
            receipt_data["date"] = expense_date.strftime("%d.%m.%Y")
            receipt_data["time"] = expense_date.strftime("%H:%M")
            receipt_data["date_time"] = expense_date.strftime("%d.%m.%Y, %H:%M")

            # Check daily budget limit - only for today's receipts
            daily_limit_alert = None
            if user.daily_budget_limit:
                # Use receipt date
                receipt_date = expense_date.date() if isinstance(expense_date, datetime) else expense_date
                today = date_type.today()
                
                # Only check limit if receipt is from today
                if receipt_date == today:
                    day_start = datetime.combine(today, datetime.min.time())
                    day_end = datetime.combine(today, datetime.max.time())
                    
                    today_expenses = db.query(Expense).filter(
                        Expense.user_id == user.id,
                        Expense.date >= day_start,
                        Expense.date <= day_end
                    ).all()
                    today_total = sum(exp.amount for exp in today_expenses)
                    
                    if today_total > user.daily_budget_limit:
                        daily_limit_alert = {
                            "type": "daily_limit_exceeded",
                            "message": f"⚠️ Gündəlik limit keçildi! Bu gün {today_total:.2f} AZN xərclədiniz (Limit: {user.daily_budget_limit:.2f} AZN)",
                            "today_spending": today_total,
                            "limit": user.daily_budget_limit
                        }
                    elif today_total > user.daily_budget_limit * 0.9:
                        daily_limit_alert = {
                            "type": "approaching",
                            "message": f"⚡ Gündəlik limitə yaxınlaşırsınız! Bu gün {today_total:.2f} AZN xərclədiniz (Limit: {user.daily_budget_limit:.2f} AZN)",
                            "today_spending": today_total,
                            "limit": user.daily_budget_limit
                        }

            scan_xp = gamification.award_xp(user, "scan_receipt", db)
            db.refresh(user)

            return templates.TemplateResponse("partials/receipt_result.html", {
                "request": request,
                "receipt_data": receipt_data,
                "success": True,
                "conversion_note": conversion_note,
                "xp_result": {
                    "xp_awarded": scan_xp["xp_awarded"] if scan_xp else 0,
                    "new_total": user.xp_points
                },
                "daily_limit_alert": daily_limit_alert
            })
        except Exception as save_err:
            return templates.TemplateResponse("partials/receipt_result.html", {
                "request": request,
                "receipt_data": {
                    "error": f"Yadda saxlama xətası: {save_err}"
                },
                "success": False
            })
        
    except Exception as e:
        return templates.TemplateResponse("partials/receipt_result.html", {
            "request": request,
            "receipt_data": {
                "error": f"Xəta baş verdi: {str(e)}",
                "items": [],
                "merchant": "Error",
                "total": 0.0,
                "date": datetime.now().strftime("%Y-%m-%d"),
                "suggested_category": "Error"
            },
            "success": False
        })


@app.post("/api/confirm-receipt")
async def confirm_receipt(
    request: Request,
    total: float = Form(...),
    merchant: str = Form(...),
    category: str = Form(...),
    date: str = Form(...),
    items: str = Form("[]"),
    db: Session = Depends(get_db)
):
    """Save receipt after user confirmation/correction"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    try:
        # Parse date safely - use UTC+4 (Azerbaijan time) or browser time
        try:
            if not date:
                from datetime import timezone as tz
                az_timezone = tz(timedelta(hours=4))
                expense_date = datetime.now(az_timezone).replace(tzinfo=None)
            else:
                expense_date = datetime.strptime(date, "%Y-%m-%d")
        except ValueError:
            from datetime import timezone as tz
            az_timezone = tz(timedelta(hours=4))
            expense_date = datetime.now(az_timezone).replace(tzinfo=None)
        
        # Parse items JSON
        import json
        try:
            items_list = json.loads(items) if items else []
        except:
            items_list = []
        
        # Save to database with user-confirmed/corrected amount
        expense = Expense(
            user_id=user.id,
            amount=total,
            merchant=merchant,
            category=category,
            date=expense_date,
            items=items_list
        )
        db.add(expense)
        db.commit()
        
        # Check daily budget limit - only for today's receipts
        daily_limit_alert = None
        if user.daily_budget_limit:
            # Parse date from form
            try:
                if not date:
                    receipt_date = date_type.today()
                else:
                    receipt_date = datetime.strptime(date, "%Y-%m-%d").date()
            except:
                receipt_date = date_type.today()
            
            today = date_type.today()
            
            # Only check limit if receipt is from today
            if receipt_date == today:
                day_start = datetime.combine(today, datetime.min.time())
                day_end = datetime.combine(today, datetime.max.time())
                
                today_expenses = db.query(Expense).filter(
                    Expense.user_id == user.id,
                    Expense.date >= day_start,
                    Expense.date <= day_end
                ).all()
                today_total = sum(exp.amount for exp in today_expenses)
                
                if today_total > user.daily_budget_limit:
                    daily_limit_alert = {
                        "type": "daily_limit_exceeded",
                        "message": f"⚠️ Gündəlik limit keçildi! Bu gün {today_total:.2f} AZN xərclədiniz (Limit: {user.daily_budget_limit:.2f} AZN)",
                        "today_spending": today_total,
                        "limit": user.daily_budget_limit
                    }
                elif today_total > user.daily_budget_limit * 0.9:
                    daily_limit_alert = {
                        "type": "approaching",
                        "message": f"⚡ Gündəlik limitə yaxınlaşırsınız! Bu gün {today_total:.2f} AZN xərclədiniz (Limit: {user.daily_budget_limit:.2f} AZN)",
                        "today_spending": today_total,
                        "limit": user.daily_budget_limit
                    }
        
        # Award XP for scanning receipt
        scan_xp = gamification.award_xp(user, "scan_receipt", db)
        db.refresh(user)
        
        # Return success result with extended timer
        return templates.TemplateResponse("partials/receipt_result.html", {
            "request": request,
            "receipt_data": {
                "total": total,
                "merchant": merchant,
                "suggested_category": category,
                "date": date,
                "items": items_list
            },
            "success": True,
            "xp_result": {
                "xp_awarded": scan_xp["xp_awarded"] if scan_xp else 0,
                "new_total": user.xp_points
            },
            "daily_limit_alert": daily_limit_alert
        })
        
    except Exception as e:
        return templates.TemplateResponse("partials/receipt_result.html", {
            "request": request,
            "receipt_data": {
                "error": f"Yadda saxlama xətası: {str(e)}"
            },
            "success": False
        })


@app.get("/profile", response_class=HTMLResponse)
async def profile_page(request: Request, db: Session = Depends(get_db)):
    """User profile page"""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    
    # Calculate stats
    total_expenses = db.query(Expense).filter(Expense.user_id == user.id).count()
    total_spent_all_time = db.query(func.sum(Expense.amount)).filter(
        Expense.user_id == user.id
    ).scalar() or 0
    total_spent_display = convert_currency(total_spent_all_time, "AZN", user.currency or "AZN")
    
    subscriptions = db.query(Expense).filter(
        Expense.user_id == user.id,
        Expense.is_subscription == True
    ).all()
    
    # Get gamification info
    level_info = gamification.get_level_info(user.xp_points)
    next_level = gamification.get_next_level_info(user.xp_points)
    
    # Get XP Logs
    xp_logs = db.query(XPLog).filter(XPLog.user_id == user.id).order_by(XPLog.created_at.desc()).limit(20).all()
    
    # Calculate breakdown
    xp_breakdown = db.query(
        XPLog.action_type, 
        func.sum(XPLog.amount)
    ).filter(
        XPLog.user_id == user.id
    ).group_by(
        XPLog.action_type
    ).all()
    
    xp_breakdown_dict = {action: amount for action, amount in xp_breakdown}

    # ===== NEW: FINANCIAL PERSONALITY DETECTION =====
    personality = detect_financial_personality(user.id, db)
    
    return templates.TemplateResponse("profile.html", {
        "request": request,
        "user": user,
        "total_expenses": total_expenses,
        "total_spent_all_time": total_spent_display,
        "subscriptions": subscriptions,
        "level_info": level_info,
        "next_level": next_level,
        "xp_logs": xp_logs,
        "xp_breakdown": xp_breakdown_dict,
        "personality": personality
    })


@app.get("/api/dashboard-data")
async def get_dashboard_data(request: Request, db: Session = Depends(get_db)):
    """API endpoint for dashboard chart data"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    # Get current month's expenses
    now = datetime.utcnow()
    month_start = datetime(now.year, now.month, 1)
    
    expenses = db.query(Expense).filter(
        Expense.user_id == user.id,
        Expense.date >= month_start
    ).all()
    
    # Category breakdown
    category_data = {}
    for exp in expenses:
        if exp.category not in category_data:
            category_data[exp.category] = 0
        category_data[exp.category] += exp.amount
    
    # Daily spending for line chart (last 30 days)
    daily_spending = {}
    for i in range(30):
        day = now - timedelta(days=i)
        day_key = day.strftime("%Y-%m-%d")
        daily_spending[day_key] = 0
    
    for exp in db.query(Expense).filter(
        Expense.user_id == user.id,
        Expense.date >= now - timedelta(days=30)
    ).all():
        day_key = exp.date.strftime("%Y-%m-%d")
        if day_key in daily_spending:
            daily_spending[day_key] += exp.amount
    
    return JSONResponse({
        "category_data": category_data,
        "daily_spending": dict(sorted(daily_spending.items()))
    })


@app.get("/heatmap", response_class=HTMLResponse)
async def heatmap_page(request: Request, db: Session = Depends(get_db)):
    """Spending heatmap page (Leaflet)"""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    expenses = db.query(Expense).filter(Expense.user_id == user.id).all()
    points = []
    for exp in expenses:
        lat, lon = pseudo_coords_for_merchant(exp.merchant)
        points.append({
            "merchant": exp.merchant,
            "amount": exp.amount,
            "category": exp.category,
            "lat": lat,
            "lon": lon
        })
    return templates.TemplateResponse("heatmap.html", {
        "request": request,
        "user": user,
        "points": points
    })


@app.get("/api/ghost-subscriptions")
async def ghost_subscriptions(request: Request, db: Session = Depends(get_db)):
    """Detect potential hidden subscriptions"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    now = datetime.utcnow()
    month_start = datetime(now.year, now.month, 1)
    expenses = db.query(Expense).filter(Expense.user_id == user.id, Expense.date >= month_start).all()

    by_merchant = {}
    for exp in expenses:
        key = (exp.merchant, round(exp.amount, 2))
        by_merchant.setdefault(key, []).append(exp)

    suspects = []
    for (merchant, amount), exps in by_merchant.items():
        if len(exps) >= 2 or any(e.is_subscription for e in exps):
            suspects.append({
                "merchant": merchant,
                "amount": amount,
                "count": len(exps),
                "category": exps[0].category,
                "latest": max(e.date for e in exps).strftime("%d.%m.%Y")
            })

    return JSONResponse({"suspects": suspects})


# ==================== NEW API ENDPOINTS ====================

@app.put("/api/expenses/{expense_id}")
async def update_expense(
    request: Request,
    expense_id: int,
    merchant: str = Form(...),
    amount: float = Form(...),
    category: str = Form(...),
    db: Session = Depends(get_db)
):
    """Update an existing expense"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    expense = db.query(Expense).filter(Expense.id == expense_id, Expense.user_id == user.id).first()
    
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    expense.merchant = merchant
    expense.amount = amount
    expense.category = category
    
    db.commit()
    db.refresh(expense)
    
    return templates.TemplateResponse("partials/transaction_row.html", {
        "request": request,
        "expense": expense
    })


@app.post("/api/voice-command")
async def voice_command_endpoint(
    request: Request,
    file: UploadFile = File(...),
    language: str = Form("az"),
    db: Session = Depends(get_db)
):
    """Process voice command and create expense"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    if not user.voice_enabled:
        return JSONResponse({"success": False, "error": "Səsli əmrlər deaktiv edilib"}, status_code=403)
    
    try:
        # Read file content
        audio_data = await file.read()
        mime_type = file.content_type
        
        # Process voice command (transcribe and parse)
        result = await voice_service.process_voice_command(audio_data, user, db, language, mime_type, save_to_db=False)
        
        if not result.get("success"):
            return JSONResponse({"success": False, "error": result.get("error")}, status_code=400)
        
        # Return confirmation dialog instead of saving immediately
        # This allows user to review and edit before final save
        return templates.TemplateResponse("partials/voice_confirmation.html", {
            "request": request,
            "transcribed_text": result["transcribed_text"],
            "expense_data": result["expense_data"]
        })
        
    except Exception as e:
        print(f"❌ Voice Command Error: {e}")
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


@app.post("/api/confirm-voice")
async def confirm_voice_expense(
    request: Request,
    amount: float = Form(...),
    merchant: str = Form(...),
    category: str = Form(...),
    db: Session = Depends(get_db)
):
    """Save voice expense after user confirmation"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    try:
        # Create expense
        expense = Expense(
            user_id=user.id,
            amount=amount,
            merchant=merchant,
            category=category,
            date=datetime.utcnow()
        )
        db.add(expense)
        db.commit()
        
        # Award XP - Fixed to 15 XP for voice commands
        xp_result = gamification.award_xp(user, "voice_command", db)
        db.refresh(user)
        
        # Always award 15 XP for voice commands
        xp_awarded = 15
        
        # Return success response - Use toast notification
        # Use window.showToast if available, otherwise create toast directly
        return HTMLResponse(content=f"""
            <script id="voice-success-script">
                // This script will be executed manually in voice_confirmation.html
                (function() {{
                    'use strict';
                    
                    // Trigger dashboard update
                    try {{
                        if (document.body) {{
                            document.body.dispatchEvent(new Event('expensesUpdated'));
                        }}
                    }} catch(e) {{
                        console.log('Event error:', e);
                    }}
                    
                    // Try to use global showToast function first
                    if (typeof window.showToast === 'function') {{
                        window.showToast(' Uğurla əlavə olundu! {amount} AZN - {merchant} (+{xp_awarded} XP)', 'success');
                        return;
                    }}
                    
                    // Fallback: Create toast manually
                    var container = document.getElementById('toast-container');
                    if (!container) {{
                        container = document.createElement('div');
                        container.id = 'toast-container';
                        container.className = 'fixed top-24 right-4 sm:right-6 z-[100] flex flex-col gap-2 pointer-events-none';
                        if (document.body) {{
                            document.body.appendChild(container);
                        }} else {{
                            setTimeout(arguments.callee, 50);
                            return;
                        }}
                    }}
                    
                    // Create toast
                    var toast = document.createElement('div');
                    toast.className = 'pointer-events-auto backdrop-blur-md text-white px-4 sm:px-6 py-3 sm:py-4 rounded-xl shadow-lg border flex items-center gap-3 transform transition-all duration-500 translate-x-full bg-green-500/90 border-green-400/50 shadow-green-500/20';
                    toast.innerHTML = '<span class="text-xl">✅</span><span class="font-medium text-sm sm:text-base">Uğurla əlavə olundu! {amount} AZN - {merchant} (+{xp_awarded} XP)</span>';
                    
                    container.appendChild(toast);
                    
                    // Animate in
                    setTimeout(function() {{
                        toast.classList.remove('translate-x-full');
                    }}, 50);
                    
                    // Remove after 3 seconds
                    setTimeout(function() {{
                        toast.classList.add('translate-x-full', 'opacity-0');
                        setTimeout(function() {{
                            if (toast && toast.parentNode) {{
                                toast.parentNode.removeChild(toast);
                            }}
                        }}, 500);
                    }}, 3000);
                }})();
            </script>
        """)
        
    except Exception as e:
        print(f"❌ Voice confirmation error: {e}")
        return HTMLResponse(content=f"""
            <div class="fixed inset-0 z-50 flex items-center justify-center bg-black/50">
                <div class="glass-card p-6 text-center max-w-sm">
                    <div class="text-6xl mb-4">❌</div>
                    <h3 class="text-xl font-bold text-white">Xəta baş verdi</h3>
                </div>
            </div>
        """, status_code=500)


@app.get("/api/export-xlsx")
async def export_xlsx(request: Request, db: Session = Depends(get_db)):
    """Generate and download Excel (XLSX) report"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        # Get current month expenses
        now = datetime.utcnow()
        month_start = datetime(now.year, now.month, 1)
        expenses = db.query(Expense).filter(
            Expense.user_id == user.id,
            Expense.date >= month_start
        ).order_by(Expense.date.desc()).all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses"
        
        # Header styling
        header_fill = PatternFill(start_color="6B46C1", end_color="6B46C1", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        # Headers
        headers = ["Tarix", "Merchant", "Kateqoriya", "Məbləğ (₼)", "Qeyd"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        total = 0
        for row, expense in enumerate(expenses, 2):
            ws.cell(row=row, column=1, value=expense.date.strftime("%d.%m.%Y %H:%M"))
            ws.cell(row=row, column=2, value=expense.merchant)
            ws.cell(row=row, column=3, value=expense.category)
            amount_cell = ws.cell(row=row, column=4, value=expense.amount)
            amount_cell.number_format = '#,##0.00'
            ws.cell(row=row, column=5, value=expense.notes or "")
            total += expense.amount
        
        # Total row
        total_row = len(expenses) + 2
        ws.cell(row=total_row, column=3, value="CƏMI:").font = Font(bold=True)
        total_cell = ws.cell(row=total_row, column=4, value=total)
        total_cell.font = Font(bold=True)
        total_cell.number_format = '#,##0.00'
        
        # Column widths
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return Response(
            content=excel_file.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=finmate_expenses_{now.strftime('%Y%m%d')}.xlsx"}
        )
    except Exception as e:
        print(f"❌ XLSX Export Error: {e}")
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


@app.get("/api/forecast")
async def get_forecast_endpoint(request: Request, db: Session = Depends(get_db)):
    """Get spending forecast for current month"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    try:
        forecast = forecast_service.get_forecast(user.id, db)
        return JSONResponse(forecast)
    except Exception as e:
        print(f"❌ Forecast Error: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/api/forecast-chart")
async def get_forecast_chart_endpoint(request: Request, db: Session = Depends(get_db)):
    """Get forecast data points for chart"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    try:
        forecast_points = forecast_service.get_chart_forecast_data(user.id, db)
        return JSONResponse({"forecast_points": forecast_points})
    except Exception as e:
        print(f"❌ Forecast Chart Error: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/api/tts")
async def text_to_speech(
    text: str = Form(...),
    language: str = Form("az")
):
    """Convert text to speech using edge-tts"""
    try:
        audio_bytes = await voice_service.generate_voice_response(text, language)
        if not audio_bytes:
            return JSONResponse({"success": False, "error": "TTS failed"}, status_code=500)
        return JSONResponse({
            "success": True,
            "audio_response": base64.b64encode(audio_bytes).decode()
        })
    except Exception as e:
        print(f"❌ TTS API Error: {e}")
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


@app.post("/api/reset-demo")
async def reset_demo_endpoint():
    """Reset database to curated demo data (manual action)"""
    try:
        reset_demo_data()
        return JSONResponse({"success": True, "message": "Demo məlumatları yeniləndi"})
    except Exception as e:
        print(f"❌ Reset Demo Error: {e}")
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


@app.post("/api/activate-trial")
async def activate_trial_endpoint(db: Session = Depends(get_db)):
    """Activate premium trial for demo user"""
    # Force reload
    try:
        user = db.query(User).filter(User.username == "demo_user").first()
        if user:
            user.is_premium = True
            db.commit()
            return JSONResponse({"success": True, "message": "Premium aktivləşdirildi! 14 gün pulsuz sınaq başladı."})
        return JSONResponse({"success": False, "error": "İstifadəçi tapılmadı"}, status_code=404)
    except Exception as e:
        print(f"❌ Trial Activation Error: {e}")
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


@app.post("/api/expense")
async def add_manual_expense(
    request: Request,
    amount: float = Form(...),
    merchant: str = Form(...),
    category: str = Form(...),
    notes: str = Form(""),
    db: Session = Depends(get_db)
):
    """Add manual expense"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    try:
        expense = Expense(
            user_id=user.id,
            amount=amount,
            merchant=merchant,
            category=category,
            notes=notes,
            date=datetime.utcnow()
        )
        db.add(expense)
        db.commit()
        
        # Check daily budget limit
        daily_limit_alert = None
        if user.daily_budget_limit:
            today = date_type.today()
            today_expenses = db.query(Expense).filter(
                Expense.user_id == user.id,
                Expense.date >= today,
                Expense.date < today + timedelta(days=1)
            ).all()
            today_total = sum(exp.amount for exp in today_expenses)
            
            if today_total > user.daily_budget_limit:
                daily_limit_alert = {
                    "type": "daily_limit_exceeded",
                    "message": f"⚠️ Gündəlik limit keçildi! Bu gün {today_total:.2f} AZN xərclədiniz (Limit: {user.daily_budget_limit:.2f} AZN)",
                    "today_spending": today_total,
                    "limit": user.daily_budget_limit
                }
        
        # Award XP
        xp_result = gamification.award_xp(user, "manual_expense", db)
        
        response_data = {
            "success": True,
            "expense_id": expense.id,
            "xp_result": xp_result
        }
        
        if daily_limit_alert:
            response_data["daily_limit_alert"] = daily_limit_alert
        
        return JSONResponse(response_data)
        
    except Exception as e:
        print(f"❌ Add Expense Error: {e}")
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


@app.delete("/api/expenses/{expense_id}")
async def delete_expense(request: Request, expense_id: int, db: Session = Depends(get_db)):
    """Delete an expense (with ownership verification)"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    try:
        expense = db.query(Expense).filter(
            Expense.id == expense_id,
            Expense.user_id == user.id  # Ensure ownership
        ).first()
        
        if not expense:
            return JSONResponse(
                {"success": False, "error": "Expense not found or unauthorized"},
                status_code=404
            )
        
        db.delete(expense)
        db.commit()
        
        # Return response with HX-Refresh header to reload page and recalculate totals
        return Response(
            status_code=200,
            headers={"HX-Refresh": "true"}
        )
    except Exception as e:
        print(f"❌ Delete Error: {e}")
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


@app.post("/api/wishlist")
async def add_wishlist_item(
    request: Request,
    title: str = Form(...),
    url: Optional[str] = Form(None),
    price: Optional[float] = Form(None),
    db: Session = Depends(get_db)
):
    """Add wish to 24h impulse-control list"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    try:
        locked_until = datetime.utcnow() + timedelta(hours=24)
        wish = Wish(
            user_id=user.id,
            title=title,
            url=url,
            price=price,
            locked_until=locked_until
        )
        db.add(wish)
        db.commit()
        db.refresh(wish)
        # Refresh list
        wishes = db.query(Wish).filter(Wish.user_id == user.id).order_by(Wish.created_at.desc()).limit(10).all()
        return templates.TemplateResponse("partials/wishlist_list.html", {"request": request, "wishes": wishes})
    except Exception as e:
        print(f"❌ Wishlist Error: {e}")
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


# ==================== DREAM VAULT API ENDPOINTS ====================

@app.get("/dream-vault", response_class=HTMLResponse)
async def dream_vault_page(request: Request, db: Session = Depends(get_db)):
    """Dream Vault page - visualize and track savings goals"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    # Get all dreams for user
    dreams = db.query(Dream).filter(
        Dream.user_id == user.id,
        Dream.is_completed == False
    ).order_by(Dream.priority.desc(), Dream.created_at.desc()).all()
    
    # Get completed dreams
    completed_dreams = db.query(Dream).filter(
        Dream.user_id == user.id,
        Dream.is_completed == True
    ).order_by(Dream.updated_at.desc()).limit(5).all()
    
    # Calculate total saved across all dreams
    total_saved = sum(dream.saved_amount for dream in dreams)
    total_target = sum(dream.target_amount for dream in dreams)
    
    return templates.TemplateResponse("dream_vault.html", {
        "request": request,
        "user": user,
        "dreams": dreams,
        "completed_dreams": completed_dreams,
        "total_saved": total_saved,
        "total_target": total_target,
        "now": datetime.utcnow(),
        "date_type": date_type,
        "max": max,
        "min": min
    })


@app.post("/api/dreams")
async def create_dream(
    request: Request,
    title: str = Form(...),
    description: Optional[str] = Form(None),
    target_amount: float = Form(...),
    image_url: Optional[str] = Form(None),
    category: Optional[str] = Form(None),
    target_date: Optional[str] = Form(None),
    priority: int = Form(1),
    db: Session = Depends(get_db)
):
    """Create a new dream"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    try:
        # Parse target_date if provided
        parsed_date = None
        if target_date:
            try:
                parsed_date = datetime.strptime(target_date, "%Y-%m-%d").date()
            except ValueError:
                pass
        
        dream = Dream(
            user_id=user.id,
            title=title,
            description=description,
            target_amount=target_amount,
            saved_amount=0.0,
            image_url=image_url,
            category=category or "Digər",
            target_date=parsed_date,
            priority=min(max(priority, 1), 5)  # Clamp between 1-5
        )
        db.add(dream)
        db.commit()
        db.refresh(dream)
        
        # Award XP for creating a dream
        gamification.award_xp(user, "create_dream", db)
        db.refresh(user)
        
        # Return the new dream card only (not the whole list)
        response = templates.TemplateResponse("partials/dream_card.html", {
            "request": request,
            "dream": dream,
            "now": datetime.utcnow(),
            "date_type": date_type,
            "max": max,
            "min": min
        })
        # Trigger stats update
        response.headers["HX-Trigger"] = "dreamUpdated"
        return response
    except Exception as e:
        print(f"❌ Create Dream Error: {e}")
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


@app.put("/api/dreams/{dream_id}")
async def update_dream(
    request: Request,
    dream_id: int,
    title: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    target_amount: Optional[float] = Form(None),
    image_url: Optional[str] = Form(None),
    category: Optional[str] = Form(None),
    target_date: Optional[str] = Form(None),
    priority: Optional[int] = Form(None),
    db: Session = Depends(get_db)
):
    """Update a dream"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    dream = db.query(Dream).filter(Dream.id == dream_id, Dream.user_id == user.id).first()
    
    if not dream:
        raise HTTPException(status_code=404, detail="Dream not found")
    
    try:
        if title is not None:
            dream.title = title
        if description is not None:
            dream.description = description
        if target_amount is not None:
            dream.target_amount = target_amount
        if image_url is not None:
            dream.image_url = image_url
        if category is not None:
            dream.category = category
        if target_date is not None:
            try:
                dream.target_date = datetime.strptime(target_date, "%Y-%m-%d").date()
            except ValueError:
                pass
        if priority is not None:
            dream.priority = min(max(priority, 1), 5)
        
        dream.updated_at = datetime.utcnow()
        db.commit()
        db.refresh(dream)
        
        # Return updated dream card
        response = templates.TemplateResponse("partials/dream_card.html", {
            "request": request,
            "dream": dream,
            "now": datetime.utcnow(),
            "date_type": date_type,
            "max": max,
            "min": min
        })
        # Trigger stats update
        response.headers["HX-Trigger"] = "dreamUpdated"
        return response
    except Exception as e:
        print(f"❌ Update Dream Error: {e}")
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


@app.post("/api/dreams/{dream_id}/add-savings")
async def add_dream_savings(
    request: Request,
    dream_id: int,
    amount: float = Form(...),
    db: Session = Depends(get_db)
):
    """Add savings to a dream"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    dream = db.query(Dream).filter(Dream.id == dream_id, Dream.user_id == user.id).first()
    
    if not dream:
        raise HTTPException(status_code=404, detail="Dream not found")
    
    try:
        dream.saved_amount += amount
        
        # Check if dream is completed
        if dream.saved_amount >= dream.target_amount:
            dream.saved_amount = dream.target_amount
            dream.is_completed = True
            dream.updated_at = datetime.utcnow()
            # Award bonus XP for completing a dream
            gamification.award_xp(user, "complete_dream", db)
        
        db.commit()
        db.refresh(dream)
        
        # Return updated dream card
        response = templates.TemplateResponse("partials/dream_card.html", {
            "request": request,
            "dream": dream,
            "now": datetime.utcnow(),
            "date_type": date_type,
            "max": max,
            "min": min
        })
        # Trigger stats update
        response.headers["HX-Trigger"] = "dreamUpdated"
        return response
    except Exception as e:
        print(f"❌ Add Savings Error: {e}")
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


@app.delete("/api/dreams/{dream_id}")
async def delete_dream(request: Request, dream_id: int, db: Session = Depends(get_db)):
    """Delete a dream"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    dream = db.query(Dream).filter(Dream.id == dream_id, Dream.user_id == user.id).first()
    
    if not dream:
        return JSONResponse({"success": False, "error": "Dream not found"}, status_code=404)
    
    try:
        db.delete(dream)
        db.commit()
        # Return empty response and trigger updates
        return Response(
            status_code=200,
            headers={
                "HX-Trigger": "dreamUpdated",
                "HX-Reswap": "none"
            }
        )
    except Exception as e:
        print(f"❌ Delete Dream Error: {e}")
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


@app.get("/api/dream-stats")
async def get_dream_stats(request: Request, db: Session = Depends(get_db)):
    """Get dynamic dream statistics"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    # Get all active dreams
    dreams = db.query(Dream).filter(
        Dream.user_id == user.id,
        Dream.is_completed == False
    ).all()
    
    # Calculate totals
    total_saved = sum(dream.saved_amount for dream in dreams)
    total_target = sum(dream.target_amount for dream in dreams)
    
    return templates.TemplateResponse("partials/dream_stats.html", {
        "request": request,
        "total_saved": total_saved,
        "total_target": total_target,
        "max": max,
        "min": min
    })


@app.get("/api/dreams/{dream_id}/data")
async def get_dream_data(request: Request, dream_id: int, db: Session = Depends(get_db)):
    """Get dream data for editing"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    dream = db.query(Dream).filter(Dream.id == dream_id, Dream.user_id == user.id).first()
    
    if not dream:
        return JSONResponse({"success": False, "error": "Dream not found"}, status_code=404)
    
    return JSONResponse({
        "success": True,
        "dream": {
            "id": dream.id,
            "title": dream.title,
            "description": dream.description,
            "target_amount": dream.target_amount,
            "saved_amount": dream.saved_amount,
            "image_url": dream.image_url,
            "category": dream.category,
            "target_date": dream.target_date.strftime("%Y-%m-%d") if dream.target_date else None,
            "priority": dream.priority
        }
    })


@app.post("/api/set-budget")
async def set_budget(
    request: Request,
    monthly_budget: float = Form(...),
    db: Session = Depends(get_db)
):
    """Set user monthly budget (onboarding)"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    try:
        if monthly_budget < 100:
            return JSONResponse({"success": False, "error": "Minimum büdcə 100 AZN-dir"}, status_code=400)
        if monthly_budget > 50000:
            return JSONResponse({"success": False, "error": "Maksimum büdcə 50000 AZN-dir"}, status_code=400)
        
        # Set budget and income to the same value
        user.monthly_budget = monthly_budget
        user.monthly_income = monthly_budget
        db.commit()
        
        return JSONResponse({"success": True, "message": "Büdcə yadda saxlanıldı"})
    except Exception as e:
        print(f"❌ Set Budget Error: {e}")
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


@app.get("/api/settings")
async def get_settings(request: Request, db: Session = Depends(get_db)):
    """Get user settings"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    return JSONResponse({
        "monthly_budget": user.monthly_budget,
        "daily_budget_limit": user.daily_budget_limit,
        "preferred_language": user.preferred_language,
        "voice_enabled": user.voice_enabled,
        "readability_mode": user.readability_mode,
        "currency": user.currency,
        "login_streak": user.login_streak,
        "ai_name": user.ai_name,
        "ai_persona_mode": user.ai_persona_mode,
        "ai_attitude": user.ai_attitude,
        "ai_style": user.ai_style
    })


@app.post("/api/settings")
async def update_settings(
    request: Request,
    db: Session = Depends(get_db)
):
    """Update user settings"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")

    def parse_bool(val):
        if val is None:
            return None
        if isinstance(val, list) and val:
            val = val[-1]
        if isinstance(val, bool):
            return val
        return str(val).lower() in ["true", "1", "on", "yes", "y", "checked"]

    try:
        form = await request.form()
        def last_value(key):
            vals = form.getlist(key)
            return vals[-1] if vals else None

        monthly_budget = last_value("monthly_budget")
        daily_budget_limit = last_value("daily_budget_limit")
        preferred_language = last_value("preferred_language")
        voice_enabled = parse_bool(form.getlist("voice_enabled"))
        readability_mode = parse_bool(form.getlist("readability_mode"))
        currency = last_value("currency")
        ai_name = last_value("ai_name")
        ai_persona_mode = last_value("ai_persona_mode")
        ai_attitude = last_value("ai_attitude")
        ai_style = last_value("ai_style")

        if monthly_budget not in (None, ""):
            try:
                budget_val = float(monthly_budget)
                if budget_val < 100:
                    raise ValueError("Budget cannot be less than 100 AZN")
                if budget_val > 50000:
                    raise ValueError("Budget cannot exceed 50000 AZN")
                user.monthly_budget = budget_val
            except (ValueError, TypeError) as e:
                print(f"❌ Invalid monthly_budget: {monthly_budget}, error: {e}")
                return JSONResponse({"success": False, "error": "Yanlış büdcə dəyəri (100-50000 AZN arası)"}, status_code=400)
        
        # Handle daily_budget_limit - allow empty string to clear it
        if daily_budget_limit is not None and daily_budget_limit != "":
            try:
                daily_val = float(daily_budget_limit)
                if daily_val < 0:
                    raise ValueError("Daily limit cannot be negative")
                if daily_val > 1000:
                    raise ValueError("Daily limit cannot exceed 1000 AZN")
                user.daily_budget_limit = daily_val
            except (ValueError, TypeError) as e:
                print(f"❌ Invalid daily_budget_limit: {daily_budget_limit}, error: {e}")
                return JSONResponse({"success": False, "error": "Yanlış gündəlik limit dəyəri (0-1000 AZN arası)"}, status_code=400)
        else:
            # Empty string or None means clear the limit
            user.daily_budget_limit = None

        if preferred_language is not None:
            user.preferred_language = preferred_language
        if voice_enabled is not None:
            user.voice_enabled = voice_enabled
        if readability_mode is not None:
            user.readability_mode = readability_mode
        if currency:
            user.currency = currency.upper()

        # AI Persona Settings
        if ai_name and ai_name.strip():
            user.ai_name = ai_name.strip()
        if ai_persona_mode is not None:
            user.ai_persona_mode = ai_persona_mode
        if ai_attitude is not None:
            user.ai_attitude = ai_attitude
        if ai_style is not None:
            user.ai_style = ai_style

        db.commit()
        return JSONResponse({"success": True, "message": "Settings updated successfully"})
    except Exception as e:
        print(f"❌ Settings Update Error: {e}")
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)



@app.get("/api/notifications")
async def get_notifications(request: Request, db: Session = Depends(get_db)):
    """Generate dynamic notifications based on user data"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    notifications = []
    
    # Get current month's data
    now = datetime.utcnow()
    month_start = datetime(now.year, now.month, 1)
    expenses = db.query(Expense).filter(
        Expense.user_id == user.id,
        Expense.date >= month_start
    ).all()
    
    total_spending = sum(exp.amount for exp in expenses)
    budget_percentage = (total_spending / user.monthly_budget * 100) if user.monthly_budget > 0 else 0
    
    # Budget warning
    if budget_percentage >= 100:
        notifications.append({
            "icon": "⚠️",
            "color": "red-500",
            "message": f"Büdcə limiti keçildi! {budget_percentage:.0f}% istifadə edilib."
        })
    elif budget_percentage >= 80:
        notifications.append({
            "icon": "⚡",
            "color": "amber-500",
            "message": f"Diqqət: Büdcənin {budget_percentage:.0f}%-ni istifadə etmisən."
        })
    
    # Daily budget limit check
    if user.daily_budget_limit:
        today = date_type.today()
        today_expenses = db.query(Expense).filter(
            Expense.user_id == user.id,
            Expense.date >= today,
            Expense.date < today + timedelta(days=1)
        ).all()
        today_total = sum(exp.amount for exp in today_expenses)
        
        if today_total > user.daily_budget_limit:
            notifications.append({
                "icon": "🚨",
                "color": "red-500",
                "message": f"Gündəlik limit keçildi! Bu gün {today_total:.2f} AZN xərclədiniz (Limit: {user.daily_budget_limit:.2f} AZN)"
            })
        elif today_total >= user.daily_budget_limit * 0.9:
            notifications.append({
                "icon": "⚡",
                "color": "amber-500",
                "message": f"Gündəlik limitə yaxınlaşırsınız! Bu gün {today_total:.2f} AZN xərclədiniz (Limit: {user.daily_budget_limit:.2f} AZN)"
            })
    
    # Subscription reminder
    subscriptions = db.query(Expense).filter(
        Expense.user_id == user.id,
        Expense.is_subscription == True
    ).all()
    
    if subscriptions:
        sub_names = [sub.merchant for sub in subscriptions[:2]]
        if len(sub_names) == 1:
            notifications.append({
                "icon": "🎬",
                "color": "purple-500",
                "message": f"{sub_names[0]} abunəliyinizi yoxlayın."
            })
        elif len(sub_names) > 1:
            notifications.append({
                "icon": "💳",
                "color": "purple-500",
                "message": f"{len(subscriptions)} aktiv abunəliyiniz var."
            })
    
    # Spending trend
    last_month_start = month_start - timedelta(days=month_start.day)
    last_month_end = month_start - timedelta(days=1)
    last_month_expenses = db.query(Expense).filter(
        Expense.user_id == user.id,
        Expense.date >= last_month_start,
        Expense.date <= last_month_end
    ).all()
    
    if last_month_expenses:
        last_month_total = sum(exp.amount for exp in last_month_expenses)
        if last_month_total > 0:
            increase = ((total_spending - last_month_total) / last_month_total) * 100
            if increase > 15:
                notifications.append({
                    "icon": "📈",
                    "color": "blue-500",
                    "message": f"Keçən aya görə {increase:.0f}% çox xərcləyirsən."
                })
            elif increase < -15:
                notifications.append({
                    "icon": "🎉",
                    "color": "green-500",
                    "message": f"Afərin! Keçən aya görə {abs(increase):.0f}% az xərclədin."
                })
    
    # XP achievement notification
    if user.xp_points > 0 and user.xp_points % 100 < 20:
        next_milestone = ((user.xp_points // 100) + 1) * 100
        remaining = next_milestone - user.xp_points
        notifications.append({
            "icon": "⭐",
            "color": "yellow-500",
            "message": f"{next_milestone} XP-yə çatmağa {remaining} XP qalıb!"
        })
    
    # Positive message if no notifications
    if not notifications:
        notifications.append({
            "icon": "✅",
            "color": "green-500",
            "message": "Maliyyə vəziyyətiniz yaxşıdır!"
        })
    
    return JSONResponse({"notifications": notifications})

@app.get("/api/dashboard-updates")
async def get_dashboard_updates(request: Request, db: Session = Depends(get_db)):
    """Return updated dashboard stats and transaction list for HTMX OOB swap"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    # Calculate total spending for current month
    now = datetime.now()
    start_of_month = datetime(now.year, now.month, 1)
    
    expenses = db.query(Expense).filter(
        Expense.user_id == user.id,
        Expense.date >= start_of_month
    ).all()
    
    total_spending = sum(e.amount for e in expenses)
    
    # Category breakdown for live updates (matches main dashboard view)
    category_data = {}
    for exp in expenses:
        category_data[exp.category] = category_data.get(exp.category, 0) + exp.amount
    
    # Calculate budget percentage
    budget_percentage = 0
    if user.monthly_budget > 0:
        budget_percentage = (total_spending / user.monthly_budget) * 100
        
    # Get recent expenses by creation time so new scans show immediately
    recent_expenses = db.query(Expense).filter(
        Expense.user_id == user.id
    ).order_by(Expense.created_at.desc()).limit(10).all()
    
    return templates.TemplateResponse("partials/dashboard_updates.html", {
        "request": request,
        "user": user,
        "total_spending": total_spending,
        "budget_percentage": budget_percentage,
        "category_data": category_data,
        "recent_expenses": recent_expenses,
        "now": now,
        "min": min,
        "float": float
    })


@app.get("/settings", response_class=HTMLResponse)
async def settings_page(request: Request, db: Session = Depends(get_db)):
    """Settings page"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    return templates.TemplateResponse("settings.html", {
        "request": request,
        "user": user,
        "min": min,
        "max": max
    })
    """User settings page"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    return templates.TemplateResponse("settings.html", {
        "request": request,
        "user": user
    })


@app.get("/api/export-pdf")
async def export_pdf(
    request: Request,
    month: Optional[int] = None,
    year: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """Export monthly report as PDF"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    
    try:
        # Default to current month/year if not specified
        now = datetime.utcnow()
        target_month = month if month else now.month
        target_year = year if year else now.year
        
        # Generate PDF
        pdf_bytes = pdf_generator.generate_monthly_report(user.id, target_month, target_year, db)
        
        # Create filename
        filename = f"finmate-report-{target_year}-{target_month:02d}.pdf"
        
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        print(f"❌ PDF Export Error: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8200, reload=True)
