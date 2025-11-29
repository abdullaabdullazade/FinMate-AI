"""Export routes"""

from fastapi import Request, Depends, HTTPException
from fastapi.responses import JSONResponse, Response
from sqlalchemy.orm import Session
from datetime import datetime
from typing import Optional
from database import get_db
from models import Expense
from config import app
from utils.auth import get_current_user
from pdf_generator import pdf_generator


@app.get("/api/export-pdf")
async def export_pdf(
    request: Request,
    month: Optional[int] = None,
    year: Optional[int] = None,
    db: Session = Depends(get_db),
):
    """Export monthly report as PDF"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")

    try:
        # Default to current month/year if not specified
        now = datetime.utcnow()
        target_month = month if month else now.month
        target_year = year if year else now.year

        # Generate PDF
        pdf_bytes = pdf_generator.generate_monthly_report(
            user.id, target_month, target_year, db
        )

        # Create filename
        filename = f"finmate-report-{target_year}-{target_month:02d}.pdf"

        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except Exception as e:
        print(f"❌ PDF Export Error: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)


@app.get("/api/export-xlsx")
async def export_xlsx(request: Request, db: Session = Depends(get_db)):
    """Generate and download Excel (XLSX) report"""
    user = get_current_user(request, db)
    if not user:
        raise HTTPException(status_code=401, detail="Authentication required")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO

        # Get current month expenses
        now = datetime.utcnow()
        month_start = datetime(now.year, now.month, 1)
        expenses = (
            db.query(Expense)
            .filter(Expense.user_id == user.id, Expense.date >= month_start)
            .order_by(Expense.date.desc())
            .all()
        )

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses"

        # Header styling
        header_fill = PatternFill(
            start_color="6B46C1", end_color="6B46C1", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True, size=12)

        # Headers
        headers = ["Tarix", "Merchant", "Kateqoriya", "Məbləğ (₼)", "Qeyd"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        total = 0
        for row, expense in enumerate(expenses, 2):
            ws.cell(row=row, column=1, value=expense.date.strftime("%d.%m.%Y %H:%M"))
            ws.cell(row=row, column=2, value=expense.merchant)
            ws.cell(row=row, column=3, value=expense.category)
            amount_cell = ws.cell(row=row, column=4, value=expense.amount)
            amount_cell.number_format = "#,##0.00"
            ws.cell(row=row, column=5, value=expense.notes or "")
            total += expense.amount

        # Total row
        total_row = len(expenses) + 2
        ws.cell(row=total_row, column=3, value="CƏMI:").font = Font(bold=True)
        total_cell = ws.cell(row=total_row, column=4, value=total)
        total_cell.font = Font(bold=True)
        total_cell.number_format = "#,##0.00"

        # Column widths
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 30

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return Response(
            content=excel_file.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=finmate_expenses_{now.strftime('%Y%m%d')}.xlsx"
            },
        )
    except Exception as e:
        print(f"❌ XLSX Export Error: {e}")
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)
